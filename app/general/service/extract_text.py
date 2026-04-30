# TODO:
#  - remove unneeded whitespace (e.g. multiple consecutive spaces)
#  - remove unprintable characters, or replacing them with some symbol
#    character so that excerpts look better.
#  - consider removing a few too common words, like single digits "1", etc.
#    or maybe anything that occurs too frequently in the full-text index that
#    could cause a full-table scan.
#  - consider multilingual stemming to enhance chances of success in a multi-
#    lingual setup... hard!


class GetTextError(Exception):
    pass


def pdf_to_text(pdf):
    # imports postponed, as they will normally not be needed frequently
    from pypdf import PdfReader
    from pypdf.errors import PdfStreamError

    text_list = []
    try:
        for page in PdfReader(pdf).pages:
            text_list.append(page.extract_text())
    except PdfStreamError:
        raise GetTextError("The uploaded PDF file is corrupted or not fully downloaded.")
    except Exception:
        raise GetTextError("Error during text extraction from PDF file.")

    return " ".join(text_list)


def excel_to_text(excel_file):
    # imports postponed, as they will normally not be needed frequently
    from openpyxl import load_workbook

    text_list = []
    try:
        workbook = load_workbook(excel_file, read_only=True, data_only=True)
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                text_list.extend(str(value) for value in row if value is not None)
    except Exception:
        raise GetTextError("Error during text extraction from Excel file.")

    return " ".join(text_list)
